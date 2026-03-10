# ============================================================
# Export des résultats dans différents formats
# ============================================================

import csv
import json
from datetime import datetime


def export_fasta(seq, filepath, header='BioSeqLab_sequence'):
    """Export de la séquence au format FASTA."""
    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(f'>{header} length={len(seq)} date={datetime.now().strftime("%Y-%m-%d")}\n')
        for i in range(0, len(seq), 60):
            f.write(seq[i:i+60] + '\n')


def export_csv(results, filepath):
    """Export de tous les résultats en CSV."""
    with open(filepath, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)

        # ORFs
        writer.writerow(['=== ORFs ==='])
        writer.writerow(['Frame', 'Debut', 'Fin', 'Longueur_pb', 'Brin', 'Nb_AA', 'Codon_Stop'])
        for orf in results.get('orfs', []):
            writer.writerow([
                orf['frame'], orf['start']+1, orf['end'],
                orf['length'], orf['strand'], orf['num_aa'], orf['stop_codon']
            ])

        writer.writerow([])

        # Promoteurs
        writer.writerow(['=== PROMOTEURS ==='])
        writer.writerow(['Pos_-35', 'Seq_-35', 'MM_-35', 'Pos_-10', 'Seq_-10', 'MM_-10', 'Espacement', 'Qualite'])
        for p in results.get('promoters', []):
            writer.writerow([
                p['pos35']+1, p['seq35'], p['dist35'],
                p['pos10']+1, p['seq10'], p['dist10'],
                p['spacing'], p['quality']
            ])

        writer.writerow([])

        # Shine-Dalgarno
        writer.writerow(['=== SHINE-DALGARNO ==='])
        writer.writerow(['Position_SD', 'Sequence_SD', 'Position_ATG', 'Espacement', 'Mismatches', 'Qualite'])
        for sd in results.get('sd_sites', []):
            writer.writerow([
                sd['pos']+1, sd['seq'], sd['atg_pos']+1,
                sd['spacing'], sd['dist'], sd['quality']
            ])

        writer.writerow([])

        # Terminateurs
        writer.writerow(['=== TERMINATEURS ==='])
        writer.writerow(['Position', 'Fin', 'Bras1', 'Boucle', 'Bras2', 'PolyT'])
        for t in results.get('terminators', []):
            writer.writerow([t['pos']+1, t['end'], t['arm1'], t['loop'], t['arm2'], t['poly_t']])

        writer.writerow([])

        # Sites de restriction
        writer.writerow(['=== SITES DE RESTRICTION ==='])
        writer.writerow(['Enzyme', 'Site', 'Nombre', 'Positions'])
        for enzyme, info in results.get('restriction', {}).items():
            writer.writerow([enzyme, info['site'], info['count'], ', '.join(map(str, info['positions']))])


def export_json(results, filepath):
    """Export complet de tous les résultats en JSON."""
    export_data = {
        'metadata': {
            'date':   datetime.now().isoformat(),
            'length': results.get('seq_length', 0),
            'gc_pct': results.get('gc_pct', 0),
        },
        'orfs': [
            {k: v for k, v in orf.items() if k != 'seq'}   # on exclut la séquence brute pour alléger
            for orf in results.get('orfs', [])
        ],
        'promoters':   results.get('promoters', []),
        'sd_sites':    results.get('sd_sites', []),
        'terminators': [
            {k: v for k, v in t.items() if k != 'seq'}
            for t in results.get('terminators', [])
        ],
        'restriction': results.get('restriction', {}),
        'stats':       results.get('stats', {}),
    }
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(export_data, f, indent=2, ensure_ascii=False)


def export_txt_report(results, seq, filepath):
    """Génère un rapport texte lisible et complet."""
    lines = []
    sep = '=' * 60
    sub = '-' * 40

    lines.append(sep)
    lines.append('  RAPPORT D\'ANALYSE ADN — BioSeq Lab')
    lines.append(f'  Date : {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    lines.append(sep)

    stats = results.get('stats', {})
    lines.append(f'\nSEQUENCE')
    lines.append(sub)
    lines.append(f'  Longueur   : {len(seq)} pb')
    lines.append(f'  Teneur GC  : {stats.get("gc", 0):.1f}%')
    lines.append(f'  Teneur AT  : {stats.get("at", 0):.1f}%')
    lines.append(f'  Composition: A={stats.get("counts", {}).get("A",0)}  '
                 f'T={stats.get("counts", {}).get("T",0)}  '
                 f'G={stats.get("counts", {}).get("G",0)}  '
                 f'C={stats.get("counts", {}).get("C",0)}')

    orfs = results.get('orfs', [])
    lines.append(f'\nORFS DETECTES ({len(orfs)})')
    lines.append(sub)
    for i, orf in enumerate(orfs[:15], 1):
        marker = ' ★' if i == 1 else ''
        lines.append(f'  [{i:2d}] Frame {orf["frame"]:+d} | '
                     f'{orf["start"]+1:>6}–{orf["end"]:<6} | '
                     f'{orf["length"]:>5} pb | '
                     f'{orf["num_aa"]:>4} aa | '
                     f'Stop: {orf["stop_codon"]}{marker}')

    promoters = results.get('promoters', [])
    lines.append(f'\nPROMOTEURS ({len(promoters)})')
    lines.append(sub)
    for i, p in enumerate(promoters, 1):
        lines.append(f'  [{i}] Box-35: {p["seq35"]} (pos.{p["pos35"]+1}) | '
                     f'Box-10: {p["seq10"]} (pos.{p["pos10"]+1}) | '
                     f'Espacement: {p["spacing"]} pb | {p["quality"]}')

    sd_sites = results.get('sd_sites', [])
    lines.append(f'\nSHINE-DALGARNO ({len(sd_sites)})')
    lines.append(sub)
    for i, sd in enumerate(sd_sites[:15], 1):
        lines.append(f'  [{i}] {sd["seq"]} | pos.{sd["pos"]+1} | '
                     f'ATG: pos.{sd["atg_pos"]+1} | '
                     f'Espacement: {sd["spacing"]} pb | {sd["quality"]}')

    terminators = results.get('terminators', [])
    lines.append(f'\nTERMINATEURS ({len(terminators)})')
    lines.append(sub)
    for i, t in enumerate(terminators[:15], 1):
        lines.append(f'  [{i}] {t["arm1"]}({t["loop"]}){t["arm2"]} + polyT:{t["poly_t"]} | '
                     f'pos.{t["pos"]+1}–{t["end"]}')

    restriction = results.get('restriction', {})
    lines.append(f'\nSITES DE RESTRICTION ({len(restriction)} enzymes)')
    lines.append(sub)
    for enzyme, info in restriction.items():
        lines.append(f'  {enzyme:<10} {info["site"]} | {info["count"]}x | pos: {info["positions"][:5]}')

    lines.append(f'\n{sep}')
    lines.append('  Rapport généré par BioSeq Lab — DNA Analyzer')
    lines.append(sep)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))


def export_excel(results, seq, filepath):
    """Export complet dans un fichier Excel multi-onglets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl requis : pip install openpyxl")

    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    header_fill  = PatternFill('solid', fgColor='1A56A0')
    alt_fill     = PatternFill('solid', fgColor='F1F5F9')
    center       = Alignment(horizontal='center')
    thin         = Side(style='thin', color='CBD5E1')
    cell_border  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, col_widths=None):
        for j, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=j, value=h)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = center
            cell.border    = cell_border
        if col_widths:
            for j, w in enumerate(col_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w

    def style_row(ws, row_idx, values):
        fill = alt_fill if row_idx % 2 == 0 else None
        for j, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=j, value=val)
            cell.border = cell_border
            cell.alignment = Alignment(horizontal='center')
            if fill:
                cell.fill = fill

    # --- Onglet ORFs ---
    ws = wb.active
    ws.title = 'ORFs'
    style_header(ws, ['#', 'Frame', 'Début', 'Fin', 'Longueur (pb)', 'Nb AA', 'Brin', 'Codon Stop'],
                 [5, 8, 10, 10, 14, 8, 8, 12])
    for i, orf in enumerate(results.get('orfs', []), 2):
        style_row(ws, i, [
            i-1, f"+{orf['frame']}", orf['start']+1, orf['end'],
            orf['length'], orf['num_aa'], orf['strand'], orf['stop_codon']
        ])

    # --- Onglet Promoteurs ---
    ws2 = wb.create_sheet('Promoteurs')
    style_header(ws2, ['#', 'Pos -35', 'Seq -35', 'MM -35', 'Pos -10', 'Seq -10', 'MM -10', 'Espacement', 'Qualité'],
                 [5, 10, 10, 8, 10, 10, 8, 12, 16])
    for i, p in enumerate(results.get('promoters', []), 2):
        style_row(ws2, i, [
            i-1, p['pos35']+1, p['seq35'], p['dist35'],
            p['pos10']+1, p['seq10'], p['dist10'],
            p['spacing'], p['quality']
        ])

    # --- Onglet Shine-Dalgarno ---
    ws3 = wb.create_sheet('Shine-Dalgarno')
    style_header(ws3, ['#', 'Position SD', 'Séquence SD', 'Position ATG', 'Espacement', 'Mismatches', 'Qualité'],
                 [5, 12, 14, 14, 12, 12, 12])
    for i, sd in enumerate(results.get('sd_sites', []), 2):
        style_row(ws3, i, [
            i-1, sd['pos']+1, sd['seq'], sd['atg_pos']+1,
            sd['spacing'], sd['dist'], sd['quality']
        ])

    # --- Onglet Terminateurs ---
    ws4 = wb.create_sheet('Terminateurs')
    style_header(ws4, ['#', 'Position', 'Fin', 'Bras 1', 'Boucle', 'Bras 2', 'PolyT', 'Longueur tige'],
                 [5, 10, 10, 12, 10, 12, 10, 14])
    for i, t in enumerate(results.get('terminators', []), 2):
        style_row(ws4, i, [
            i-1, t['pos']+1, t['end'], t['arm1'],
            t['loop'], t['arm2'], t['poly_t'], t['stem_len']
        ])

    # --- Onglet Restriction ---
    ws5 = wb.create_sheet('Restriction')
    style_header(ws5, ['Enzyme', 'Site', 'Nombre de coupures', 'Positions'],
                 [12, 12, 20, 50])
    for i, (enzyme, info) in enumerate(results.get('restriction', {}).items(), 2):
        style_row(ws5, i, [
            enzyme, info['site'], info['count'],
            ', '.join(map(str, info['positions']))
        ])

    # --- Onglet Statistiques ---
    ws6 = wb.create_sheet('Statistiques')
    stats = results.get('stats', {})
    data = [
        ['Longueur totale', f"{stats.get('length', len(seq))} pb"],
        ['Teneur en GC', f"{stats.get('gc', 0):.2f}%"],
        ['Teneur en AT', f"{stats.get('at', 0):.2f}%"],
        ['Nb de A', stats.get('counts', {}).get('A', 0)],
        ['Nb de T', stats.get('counts', {}).get('T', 0)],
        ['Nb de G', stats.get('counts', {}).get('G', 0)],
        ['Nb de C', stats.get('counts', {}).get('C', 0)],
    ]
    style_header(ws6, ['Paramètre', 'Valeur'], [25, 20])
    for i, row in enumerate(data, 2):
        style_row(ws6, i, row)

    wb.save(filepath)
